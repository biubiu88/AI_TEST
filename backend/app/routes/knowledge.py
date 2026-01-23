"""
知识库管理路由
"""
import io
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app import db
from app.models import Knowledge

knowledge_bp = Blueprint('knowledge', __name__)


@knowledge_bp.route('', methods=['GET'])
def get_knowledges():
    """获取知识库列表"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    keyword = request.args.get('keyword', '')
    category = request.args.get('category', '')
    is_active = request.args.get('is_active', '')
    
    query = Knowledge.query
    
    if keyword:
        query = query.filter(
            db.or_(
                Knowledge.name.contains(keyword),
                Knowledge.description.contains(keyword)
            )
        )
    if category:
        query = query.filter(Knowledge.category == category)
    if is_active:
        query = query.filter(Knowledge.is_active == (is_active == 'true'))
    
    query = query.order_by(Knowledge.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    return jsonify({
        'code': 0,
        'message': 'success',
        'data': {
            'list': [item.to_dict() for item in pagination.items],
            'total': pagination.total,
            'page': page,
            'per_page': per_page
        }
    })


@knowledge_bp.route('/all', methods=['GET'])
def get_all_knowledges():
    """获取所有启用的知识库（用于下拉选择）"""
    knowledges = Knowledge.query.filter(Knowledge.is_active == True).order_by(
        Knowledge.name
    ).all()
    
    return jsonify({
        'code': 0,
        'message': 'success',
        'data': [item.to_dict() for item in knowledges]
    })


@knowledge_bp.route('/batch', methods=['GET'])
def get_knowledges_by_ids():
    """根据ID列表获取知识库内容"""
    ids_str = request.args.get('ids', '')
    if not ids_str:
        return jsonify({
            'code': 0,
            'message': 'success',
            'data': []
        })
    
    ids = [int(id) for id in ids_str.split(',') if id.strip().isdigit()]
    knowledges = Knowledge.query.filter(
        Knowledge.id.in_(ids),
        Knowledge.is_active == True
    ).all()
    
    return jsonify({
        'code': 0,
        'message': 'success',
        'data': [item.to_dict() for item in knowledges]
    })


@knowledge_bp.route('/export', methods=['GET'])
def export_knowledges():
    """导出知识库为Excel"""
    # 获取筛选参数
    ids = request.args.get('ids', '')  # 支持导出指定ID
    
    if ids:
        id_list = [int(i) for i in ids.split(',') if i.strip().isdigit()]
        knowledges = Knowledge.query.filter(Knowledge.id.in_(id_list)).all()
    else:
        knowledges = Knowledge.query.all()
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "知识库"
    
    # 设置标题行样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 定义列标题
    headers = ['ID', '名称', '描述', '内容', '分类', '文件类型', '是否启用', '创建时间']
    
    # 设置列宽
    column_widths = [8, 25, 30, 80, 15, 12, 12, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 写入标题行
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 分类中文映射
    category_map = {'general': '通用', 'domain': '领域知识', 'api': 'API文档', 'ui': 'UI规范', 'database': '数据库'}
    file_type_map = {'text': '纯文本', 'markdown': 'Markdown', 'json': 'JSON'}
    
    # 写入数据行
    for row, k in enumerate(knowledges, 2):
        data = [
            k.id,
            k.name,
            k.description or '',
            k.content,
            category_map.get(k.category, k.category),
            file_type_map.get(k.file_type, k.file_type),
            '是' if k.is_active else '否',
            k.created_at.strftime('%Y-%m-%d %H:%M:%S') if k.created_at else ''
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    filename = f"知识库_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@knowledge_bp.route('/template', methods=['GET'])
def download_template():
    """下载导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "知识库导入模板"
    
    # 设置标题行样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 定义列标题（必填字段加*标记）
    headers = ['名称*', '内容*', '描述', '分类', '文件类型', '是否启用']
    
    # 设置列宽
    column_widths = [25, 80, 30, 20, 15, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 写入标题行
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入说明行
    instructions = [
        '填写知识库名称',
        '填写知识库内容',
        '可选，填写描述',
        '通用/领域知识/API文档/UI规范/数据库',
        '纯文本/Markdown/JSON',
        '是/否'
    ]
    note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for col, instruction in enumerate(instructions, 1):
        cell = ws.cell(row=2, column=col, value=instruction)
        cell.fill = note_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # 写入示例数据
    example_data = [
        ['登录功能测试知识', '登录功能要求：\n1.支持用户名+密码登录\n2.支持手机验证码登录\n3.密码错误5次锁定账户', '用户登录相关的业务规则', '领域知识', '纯文本', '是'],
        ['API接口规范', '## 用户接口\n\n### POST /api/user/login\n请求参数：username, password\n返回：token, user_info', 'API接口文档', 'API文档', 'Markdown', '是']
    ]
    for row, data in enumerate(example_data, 3):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='知识库导入模板.xlsx'
    )


@knowledge_bp.route('/import', methods=['POST'])
def import_knowledges():
    """导入知识库"""
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请上传文件'}), 400
    
    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'code': 400, 'message': '请上传Excel文件(.xlsx或.xls)'}), 400
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        # 分类和文件类型映射
        category_map = {'通用': 'general', '领域知识': 'domain', 'API文档': 'api', 'UI规范': 'ui', '数据库': 'database'}
        file_type_map = {'纯文本': 'text', 'Markdown': 'markdown', 'JSON': 'json'}
        
        created_count = 0
        updated_count = 0
        errors = []
        
        # 从第3行开始读取数据（跳过标题行和说明行）
        for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
            # 跳过空行
            if not row or not any(row):
                continue
            
            name = row[0]
            content = row[1]
            description = row[2] or ''
            category = row[3] or '通用'
            file_type = row[4] if len(row) > 4 else '纯文本'
            is_active = row[5] if len(row) > 5 else '是'
            
            # 验证必填字段
            if not name or not content:
                errors.append(f'第{row_num}行: 缺少必填字段(名称或内容)')
                continue
            
            # 转换分类、文件类型和状态
            category_val = category_map.get(str(category), 'general')
            file_type_val = file_type_map.get(str(file_type) if file_type else '纯文本', 'text')
            is_active_val = str(is_active) in ['是', 'True', 'true', '1', 'yes', 'Yes']
            
            # 检查是否已存在同名知识库
            existing = Knowledge.query.filter(Knowledge.name == str(name)).first()
            if existing:
                # 更新现有记录
                existing.content = str(content)
                existing.description = str(description) if description else ''
                existing.category = category_val
                existing.file_type = file_type_val
                existing.is_active = is_active_val
                updated_count += 1
            else:
                # 创建新记录
                knowledge = Knowledge(
                    name=str(name),
                    content=str(content),
                    description=str(description) if description else '',
                    category=category_val,
                    file_type=file_type_val,
                    is_active=is_active_val
                )
                db.session.add(knowledge)
                created_count += 1
        
        db.session.commit()
        
        result = {
            'code': 0,
            'message': f'导入成功，新增 {created_count} 条，更新 {updated_count} 条',
            'data': {
                'success_count': created_count + updated_count,
                'created_count': created_count,
                'updated_count': updated_count,
                'errors': errors
            }
        }
        
        if errors:
            result['message'] += f'，{len(errors)} 条失败'
        
        return jsonify(result)
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'导入失败: {str(e)}'}), 500


@knowledge_bp.route('/<int:knowledge_id>', methods=['GET'])
def get_knowledge(knowledge_id):
    """获取知识库详情"""
    knowledge = Knowledge.query.get_or_404(knowledge_id)
    return jsonify({
        'code': 0,
        'message': 'success',
        'data': knowledge.to_dict()
    })


@knowledge_bp.route('', methods=['POST'])
def create_knowledge():
    """创建知识库"""
    data = request.get_json()
    
    required_fields = ['name', 'content']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'code': 400, 'message': f'{field}不能为空'}), 400
    
    knowledge = Knowledge(
        name=data['name'],
        content=data['content'],
        description=data.get('description', ''),
        category=data.get('category', 'general'),
        file_type=data.get('file_type', 'text'),
        is_active=data.get('is_active', True)
    )
    
    db.session.add(knowledge)
    db.session.commit()
    
    return jsonify({
        'code': 0,
        'message': '创建成功',
        'data': knowledge.to_dict()
    })


@knowledge_bp.route('/<int:knowledge_id>', methods=['PUT'])
def update_knowledge(knowledge_id):
    """更新知识库"""
    knowledge = Knowledge.query.get_or_404(knowledge_id)
    data = request.get_json()
    
    if 'name' in data:
        knowledge.name = data['name']
    if 'content' in data:
        knowledge.content = data['content']
    if 'description' in data:
        knowledge.description = data['description']
    if 'category' in data:
        knowledge.category = data['category']
    if 'file_type' in data:
        knowledge.file_type = data['file_type']
    if 'is_active' in data:
        knowledge.is_active = data['is_active']
    
    db.session.commit()
    
    return jsonify({
        'code': 0,
        'message': '更新成功',
        'data': knowledge.to_dict()
    })


@knowledge_bp.route('/<int:knowledge_id>', methods=['DELETE'])
def delete_knowledge(knowledge_id):
    """删除知识库"""
    knowledge = Knowledge.query.get_or_404(knowledge_id)
    
    db.session.delete(knowledge)
    db.session.commit()
    
    return jsonify({
        'code': 0,
        'message': '删除成功'
    })
