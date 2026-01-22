"""
测试用例路由
"""
import io
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app import db
from app.models import TestCase, Requirement

testcase_bp = Blueprint('testcase', __name__)


@testcase_bp.route('', methods=['GET'])
def get_testcases():
    """获取测试用例列表"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    keyword = request.args.get('keyword', '')
    requirement_id = request.args.get('requirement_id', type=int)
    case_type = request.args.get('case_type', '')
    status = request.args.get('status', '')
    is_ai_generated = request.args.get('is_ai_generated', '')
    
    query = TestCase.query
    
    if keyword:
        query = query.filter(
            db.or_(
                TestCase.title.contains(keyword),
                TestCase.steps.contains(keyword)
            )
        )
    if requirement_id:
        query = query.filter(TestCase.requirement_id == requirement_id)
    if case_type:
        query = query.filter(TestCase.case_type == case_type)
    if status:
        query = query.filter(TestCase.status == status)
    if is_ai_generated:
        query = query.filter(TestCase.is_ai_generated == (is_ai_generated == 'true'))
    
    query = query.order_by(TestCase.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    return jsonify({
        'code': 0,
        'message': 'success',
        'data': {
            'list': [item.to_dict() for item in pagination.items],
            'total': pagination.total,
            'page': page,
            'per_page': per_page
        }
    })


@testcase_bp.route('/<int:testcase_id>', methods=['GET'])
def get_testcase(testcase_id):
    """获取测试用例详情"""
    testcase = TestCase.query.get_or_404(testcase_id)
    return jsonify({
        'code': 0,
        'message': 'success',
        'data': testcase.to_dict()
    })


@testcase_bp.route('', methods=['POST'])
def create_testcase():
    """创建测试用例"""
    data = request.get_json()

    required_fields = ['requirement_id', 'title', 'steps', 'expected_result']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'code': 400, 'message': f'{field}不能为空'}), 400

    # 验证需求是否存在
    requirement = Requirement.query.get(data['requirement_id'])
    if not requirement:
        return jsonify({'code': 404, 'message': '关联需求不存在'}), 404

    # 处理 steps 字段 - 如果是列表则转换为字符串
    steps = data['steps']
    if isinstance(steps, list):
        steps = '; '.join(steps)  # 或者使用 '\n'.join(steps)

    testcase = TestCase(
        requirement_id=data['requirement_id'],
        title=data['title'],
        precondition=data.get('precondition', ''),
        steps=steps,  # 使用处理后的字符串
        expected_result=data['expected_result'],
        case_type=data.get('case_type', 'functional'),
        priority=data.get('priority', 'medium'),
        status=data.get('status', 'pending'),
        is_ai_generated=data.get('is_ai_generated', False)
    )

    db.session.add(testcase)
    db.session.commit()

    return jsonify({
        'code': 0,
        'message': '创建成功',
        'data': testcase.to_dict()
    })


@testcase_bp.route('/batch', methods=['POST'])
def create_testcases_batch():
    """批量创建测试用例"""
    data = request.get_json()
    testcases_data = data.get('testcases', [])
    
    if not testcases_data:
        return jsonify({'code': 400, 'message': '测试用例列表不能为空'}), 400
    
    created_testcases = []
    for tc_data in testcases_data:
        # 处理 steps 字段 - 如果是列表则转换为字符串
        steps = tc_data['steps']
        if isinstance(steps, list):
            steps = '; '.join(steps)  # 或者使用 '\n'.join(steps)
        testcase = TestCase(
            requirement_id=tc_data['requirement_id'],
            title=tc_data['title'],
            precondition=tc_data.get('precondition', ''),
            steps=steps,
            expected_result=tc_data['expected_result'],
            case_type=tc_data.get('case_type', 'functional'),
            priority=tc_data.get('priority', 'medium'),
            status='pending',
            is_ai_generated=tc_data.get('is_ai_generated', False)
        )
        db.session.add(testcase)
        created_testcases.append(testcase)
    
    db.session.commit()
    
    return jsonify({
        'code': 0,
        'message': f'成功创建 {len(created_testcases)} 条测试用例',
        'data': [tc.to_dict() for tc in created_testcases]
    })


@testcase_bp.route('/<int:testcase_id>', methods=['PUT'])
def update_testcase(testcase_id):
    """更新测试用例"""
    testcase = TestCase.query.get_or_404(testcase_id)
    data = request.get_json()
    
    if 'title' in data:
        testcase.title = data['title']
    if 'precondition' in data:
        testcase.precondition = data['precondition']
    if 'steps' in data:
        testcase.steps = data['steps']
    if 'expected_result' in data:
        testcase.expected_result = data['expected_result']
    if 'case_type' in data:
        testcase.case_type = data['case_type']
    if 'priority' in data:
        testcase.priority = data['priority']
    if 'status' in data:
        testcase.status = data['status']
    
    db.session.commit()
    
    return jsonify({
        'code': 0,
        'message': '更新成功',
        'data': testcase.to_dict()
    })


@testcase_bp.route('/<int:testcase_id>', methods=['DELETE'])
def delete_testcase(testcase_id):
    """删除测试用例"""
    testcase = TestCase.query.get_or_404(testcase_id)
    db.session.delete(testcase)
    db.session.commit()
    
    return jsonify({
        'code': 0,
        'message': '删除成功'
    })


@testcase_bp.route('/stats', methods=['GET'])
def get_stats():
    """获取测试用例统计"""
    total = TestCase.query.count()
    passed = TestCase.query.filter_by(status='passed').count()
    failed = TestCase.query.filter_by(status='failed').count()
    pending = TestCase.query.filter_by(status='pending').count()
    blocked = TestCase.query.filter_by(status='blocked').count()
    ai_generated = TestCase.query.filter_by(is_ai_generated=True).count()
    
    return jsonify({
        'code': 0,
        'message': 'success',
        'data': {
            'total': total,
            'passed': passed,
            'failed': failed,
            'pending': pending,
            'blocked': blocked,
            'ai_generated': ai_generated,
            'pass_rate': round(passed / total * 100, 2) if total > 0 else 0
        }
    })


@testcase_bp.route('/export', methods=['GET'])
def export_testcases():
    """导出测试用例为Excel"""
    # 获取筛选参数
    keyword = request.args.get('keyword', '')
    requirement_id = request.args.get('requirement_id', type=int)
    case_type = request.args.get('case_type', '')
    status = request.args.get('status', '')
    is_ai_generated = request.args.get('is_ai_generated', '')
    
    query = TestCase.query
    
    if keyword:
        query = query.filter(
            db.or_(
                TestCase.title.contains(keyword),
                TestCase.steps.contains(keyword)
            )
        )
    if requirement_id:
        query = query.filter(TestCase.requirement_id == requirement_id)
    if case_type:
        query = query.filter(TestCase.case_type == case_type)
    if status:
        query = query.filter(TestCase.status == status)
    if is_ai_generated:
        query = query.filter(TestCase.is_ai_generated == (is_ai_generated == 'true'))
    
    testcases = query.order_by(TestCase.created_at.desc()).all()
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    
    # 设置标题行样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 定义列标题
    headers = ['ID', '关联需求', '用例标题', '前置条件', '测试步骤', '预期结果', '类型', '优先级', '状态', '来源', '创建时间']
    
    # 设置列宽
    column_widths = [8, 25, 35, 25, 50, 40, 10, 10, 10, 10, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 写入标题行
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 类型和状态的中文映射
    type_map = {'functional': '功能测试', 'boundary': '边界测试', 'exception': '异常测试', 'performance': '性能测试'}
    priority_map = {'high': '高', 'medium': '中', 'low': '低'}
    status_map = {'pending': '待执行', 'passed': '通过', 'failed': '失败', 'blocked': '阻塞'}
    
    # 写入数据行
    for row, tc in enumerate(testcases, 2):
        data = [
            tc.id,
            tc.requirement.title if tc.requirement else '',
            tc.title,
            tc.precondition or '',
            tc.steps,
            tc.expected_result,
            type_map.get(tc.case_type, tc.case_type),
            priority_map.get(tc.priority, tc.priority),
            status_map.get(tc.status, tc.status),
            'AI生成' if tc.is_ai_generated else '手动创建',
            tc.created_at.strftime('%Y-%m-%d %H:%M:%S') if tc.created_at else ''
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    filename = f"测试用例_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@testcase_bp.route('/template', methods=['GET'])
def download_template():
    """下载导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例导入模板"
    
    # 设置标题行样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 定义列标题（必填字段加*标记）
    headers = ['关联需求ID*', '用例标题*', '前置条件', '测试步骤*', '预期结果*', '类型', '优先级']
    
    # 设置列宽
    column_widths = [15, 35, 30, 50, 40, 15, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 写入标题行
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入说明行
    instructions = [
        '填写需求ID数字',
        '填写用例标题',
        '可选，填写前置条件',
        '填写测试步骤',
        '填写预期结果',
        '功能测试/边界测试/异常测试/性能测试',
        '高/中/低'
    ]
    note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for col, instruction in enumerate(instructions, 1):
        cell = ws.cell(row=2, column=col, value=instruction)
        cell.fill = note_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # 写入示例数据
    example_data = [
        [1, '测试登录功能-正常登录', '用户已注册账户', '1.打开登录页面\n2.输入用户名和密码\n3.点击登录按钮', '登录成功，跳转到首页', '功能测试', '高'],
        [1, '测试登录功能-密码错误', '用户已注册账户', '1.打开登录页面\n2.输入用户名和错误密码\n3.点击登录按钮', '显示密码错误提示', '异常测试', '中']
    ]
    for row, data in enumerate(example_data, 3):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='测试用例导入模板.xlsx'
    )


@testcase_bp.route('/import', methods=['POST'])
def import_testcases():
    """导入测试用例"""
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请上传文件'}), 400
    
    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'code': 400, 'message': '请上传Excel文件(.xlsx或.xls)'}), 400
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        # 类型和优先级的映射
        type_map = {'功能测试': 'functional', '边界测试': 'boundary', '异常测试': 'exception', '性能测试': 'performance'}
        priority_map = {'高': 'high', '中': 'medium', '低': 'low'}
        
        created_count = 0
        errors = []
        
        # 从第3行开始读取数据（跳过标题行和说明行）
        for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
            # 跳过空行
            if not row or not any(row):
                continue
            
            requirement_id = row[0]
            title = row[1]
            precondition = row[2] or ''
            steps = row[3]
            expected_result = row[4]
            case_type = row[5] or '功能测试'
            priority = row[6] or '中'
            
            # 验证必填字段
            if not requirement_id or not title or not steps or not expected_result:
                errors.append(f'第{row_num}行: 缺少必填字段')
                continue
            
            # 验证需求ID
            try:
                requirement_id = int(requirement_id)
            except (ValueError, TypeError):
                errors.append(f'第{row_num}行: 需求ID必须是数字')
                continue
            
            requirement = Requirement.query.get(requirement_id)
            if not requirement:
                errors.append(f'第{row_num}行: 需求ID {requirement_id} 不存在')
                continue
            
            # 转换类型和优先级
            case_type_val = type_map.get(case_type, 'functional')
            priority_val = priority_map.get(priority, 'medium')
            
            # 创建测试用例
            testcase = TestCase(
                requirement_id=requirement_id,
                title=str(title),
                precondition=str(precondition) if precondition else '',
                steps=str(steps),
                expected_result=str(expected_result),
                case_type=case_type_val,
                priority=priority_val,
                status='pending',
                is_ai_generated=False
            )
            db.session.add(testcase)
            created_count += 1
        
        db.session.commit()
        
        result = {
            'code': 0,
            'message': f'成功导入 {created_count} 条测试用例',
            'data': {
                'success_count': created_count,
                'errors': errors
            }
        }
        
        if errors:
            result['message'] += f'，{len(errors)} 条失败'
        
        return jsonify(result)
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'导入失败: {str(e)}'}), 500
