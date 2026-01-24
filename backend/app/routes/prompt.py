"""
提示词管理路由
"""
import io
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app import db
from app.models import Prompt
from app.middlewares import log_operation

prompt_bp = Blueprint('prompt', __name__)


@prompt_bp.route('', methods=['GET'])
@jwt_required()
@log_operation
def get_prompts():
    """获取提示词列表"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    keyword = request.args.get('keyword', '')
    category = request.args.get('category', '')
    is_active = request.args.get('is_active', '')
    
    query = Prompt.query
    
    if keyword:
        query = query.filter(
            db.or_(
                Prompt.name.contains(keyword),
                Prompt.description.contains(keyword)
            )
        )
    if category:
        query = query.filter(Prompt.category == category)
    if is_active:
        query = query.filter(Prompt.is_active == (is_active == 'true'))
    
    query = query.order_by(Prompt.is_default.desc(), Prompt.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    return jsonify({
        'code': 0,
        'message': 'success',
        'data': {
            'list': [item.to_dict() for item in pagination.items],
            'total': pagination.total,
            'page': page,
            'per_page': per_page
        }
    })


@prompt_bp.route('/all', methods=['GET'])
@jwt_required()
@log_operation
def get_all_prompts():
    """获取所有启用的提示词（用于下拉选择）"""
    prompts = Prompt.query.filter(Prompt.is_active == True).order_by(
        Prompt.is_default.desc(), Prompt.name
    ).all()
    
    return jsonify({
        'code': 0,
        'message': 'success',
        'data': [item.to_dict() for item in prompts]
    })


@prompt_bp.route('/export', methods=['GET'])
@jwt_required()
@log_operation
def export_prompts():
    """导出提示词为Excel"""
    # 获取筛选参数
    ids = request.args.get('ids', '')  # 支持导出指定ID
    
    if ids:
        id_list = [int(i) for i in ids.split(',') if i.strip().isdigit()]
        prompts = Prompt.query.filter(Prompt.id.in_(id_list)).all()
    else:
        prompts = Prompt.query.all()
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "提示词"
    
    # 设置标题行样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 定义列标题
    headers = ['ID', '名称', '描述', '内容', '分类', '是否默认', '是否启用', '创建时间']
    
    # 设置列宽
    column_widths = [8, 25, 30, 80, 15, 12, 12, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 写入标题行
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 分类中文映射
    category_map = {'general': '通用', 'functional': '功能测试', 'boundary': '边界测试', 'exception': '异常测试', 'performance': '性能测试'}
    
    # 写入数据行
    for row, p in enumerate(prompts, 2):
        data = [
            p.id,
            p.name,
            p.description or '',
            p.content,
            category_map.get(p.category, p.category),
            '是' if p.is_default else '否',
            '是' if p.is_active else '否',
            p.created_at.strftime('%Y-%m-%d %H:%M:%S') if p.created_at else ''
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    filename = f"提示词_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@prompt_bp.route('/template', methods=['GET'])
@jwt_required()
@log_operation
def download_template():
    """下载导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "提示词导入模板"
    
    # 设置标题行样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 定义列标题（必填字段加*标记）
    headers = ['名称*', '内容*', '描述', '分类', '是否启用']
    
    # 设置列宽
    column_widths = [25, 80, 30, 20, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 写入标题行
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入说明行
    instructions = [
        '填写提示词名称',
        '填写提示词内容',
        '可选，填写描述',
        '通用/功能测试/边界测试/异常测试/性能测试',
        '是/否'
    ]
    note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for col, instruction in enumerate(instructions, 1):
        cell = ws.cell(row=2, column=col, value=instruction)
        cell.fill = note_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # 写入示例数据
    example_data = [
        ['测试用例生成提示词', '你是一个专业的测试工程师，请根据用户提供的需求生成详细的测试用例...', '用于AI生成测试用例', '通用', '是'],
        ['边界测试提示词', '请重点关注输入边界值、特殊字符、极限值等边界条件...', '用于生成边界测试用例', '边界测试', '是']
    ]
    for row, data in enumerate(example_data, 3):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='提示词导入模板.xlsx'
    )


@prompt_bp.route('/import', methods=['POST'])
@jwt_required()
@log_operation
def import_prompts():
    """导入提示词"""
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请上传文件'}), 400
    
    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'code': 400, 'message': '请上传Excel文件(.xlsx或.xls)'}), 400
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        # 分类映射
        category_map = {'通用': 'general', '功能测试': 'functional', '边界测试': 'boundary', '异常测试': 'exception', '性能测试': 'performance'}
        
        created_count = 0
        updated_count = 0
        errors = []
        
        # 从第3行开始读取数据（跳过标题行和说明行）
        for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
            # 跳过空行
            if not row or not any(row):
                continue
            
            name = row[0]
            content = row[1]
            description = row[2] or ''
            category = row[3] or '通用'
            is_active = row[4] if len(row) > 4 else '是'
            
            # 验证必填字段
            if not name or not content:
                errors.append(f'第{row_num}行: 缺少必填字段(名称或内容)')
                continue
            
            # 转换分类和状态
            category_val = category_map.get(str(category), 'general')
            is_active_val = str(is_active) in ['是', 'True', 'true', '1', 'yes', 'Yes']
            
            # 检查是否已存在同名提示词
            existing = Prompt.query.filter(Prompt.name == str(name)).first()
            if existing:
                # 更新现有记录
                existing.content = str(content)
                existing.description = str(description) if description else ''
                existing.category = category_val
                existing.is_active = is_active_val
                updated_count += 1
            else:
                # 创建新记录
                prompt = Prompt(
                    name=str(name),
                    content=str(content),
                    description=str(description) if description else '',
                    category=category_val,
                    is_default=False,
                    is_active=is_active_val
                )
                db.session.add(prompt)
                created_count += 1
        
        db.session.commit()
        
        result = {
            'code': 0,
            'message': f'导入成功，新增 {created_count} 条，更新 {updated_count} 条',
            'data': {
                'success_count': created_count + updated_count,
                'created_count': created_count,
                'updated_count': updated_count,
                'errors': errors
            }
        }
        
        if errors:
            result['message'] += f'，{len(errors)} 条失败'
        
        return jsonify(result)
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'message': f'导入失败: {str(e)}'}), 500


@prompt_bp.route('/<int:prompt_id>', methods=['GET'])
@jwt_required()
@log_operation
def get_prompt(prompt_id):
    """获取提示词详情"""
    prompt = Prompt.query.get_or_404(prompt_id)
    return jsonify({
        'code': 0,
        'message': 'success',
        'data': prompt.to_dict()
    })


@prompt_bp.route('', methods=['POST'])
@jwt_required()
@log_operation
def create_prompt():
    """创建提示词"""
    data = request.get_json()
    
    required_fields = ['name', 'content']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'code': 400, 'message': f'{field}不能为空'}), 400
    
    # 如果设置为默认，先取消其他默认
    if data.get('is_default'):
        Prompt.query.filter(Prompt.is_default == True).update({'is_default': False})
    
    prompt = Prompt(
        name=data['name'],
        content=data['content'],
        description=data.get('description', ''),
        category=data.get('category', 'general'),
        is_default=data.get('is_default', False),
        is_active=data.get('is_active', True)
    )
    
    db.session.add(prompt)
    db.session.commit()
    
    return jsonify({
        'code': 0,
        'message': '创建成功',
        'data': prompt.to_dict()
    })


@prompt_bp.route('/<int:prompt_id>', methods=['PUT'])
@jwt_required()
@log_operation
def update_prompt(prompt_id):
    """更新提示词"""
    prompt = Prompt.query.get_or_404(prompt_id)
    data = request.get_json()
    
    # 如果设置为默认，先取消其他默认
    if data.get('is_default') and not prompt.is_default:
        Prompt.query.filter(Prompt.id != prompt_id, Prompt.is_default == True).update({'is_default': False})
    
    if 'name' in data:
        prompt.name = data['name']
    if 'content' in data:
        prompt.content = data['content']
    if 'description' in data:
        prompt.description = data['description']
    if 'category' in data:
        prompt.category = data['category']
    if 'is_default' in data:
        prompt.is_default = data['is_default']
    if 'is_active' in data:
        prompt.is_active = data['is_active']
    
    db.session.commit()
    
    return jsonify({
        'code': 0,
        'message': '更新成功',
        'data': prompt.to_dict()
    })


@prompt_bp.route('/<int:prompt_id>', methods=['DELETE'])
@jwt_required()
@log_operation
def delete_prompt(prompt_id):
    """删除提示词"""
    prompt = Prompt.query.get_or_404(prompt_id)
    
    db.session.delete(prompt)
    db.session.commit()
    
    return jsonify({
        'code': 0,
        'message': '删除成功'
    })


@prompt_bp.route('/<int:prompt_id>/default', methods=['PUT'])
def set_default_prompt(prompt_id):
    """设置默认提示词"""
    prompt = Prompt.query.get_or_404(prompt_id)
    
    # 取消其他默认
    Prompt.query.filter(Prompt.is_default == True).update({'is_default': False})
    
    # 设置当前为默认
    prompt.is_default = True
    db.session.commit()
    
    return jsonify({
        'code': 0,
        'message': '设置成功',
        'data': prompt.to_dict()
    })
